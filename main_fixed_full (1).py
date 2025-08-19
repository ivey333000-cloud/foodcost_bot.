import os
import re
from pathlib import Path
import pandas as pd
import openpyxl
from margin_control import setup_margin_control
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from telegram import (
    Update,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    InlineKeyboardButton,
    InlineKeyboardMarkup,

    BotCommand,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    filters,
    CallbackQueryHandler,
)

# === Постоянные команды бота (кнопка меню Telegram) ===
async def set_bot_commands(app):
    try:
        commands = [
            BotCommand("start", "Главное меню"),
            BotCommand("ttk", "Создать новую ТТК"),
            BotCommand("ttk_bulk", "Массовая загрузка ТТК"),
            BotCommand("cost", "Рассчитать себестоимость"),
            BotCommand("zakup", "Рассчитать закуп продуктов"),
            BotCommand("delete", "Удалить блюдо из ТТК"),
            BotCommand("table", "Скачать Себестоимость.xlsx"),
            BotCommand("ttk_file", "Скачать TTK.xlsx"),
            BotCommand("dishes", "Выбрать блюда для закупа"),
        ]
        await app.bot.set_my_commands(commands)
    except Exception as e:
        print(f"[WARN] set_my_commands failed: {e}")

# =====================
# Константы/файлы
# =====================
EXCEL_FILE = "Себестоимость.xlsx"
TTK_FILE = "TTK.xlsx"
TTK_SHEET_NAME = "TTK"
PAGE_SIZE = 10

# Состояния
AWAITING_TTK_NAME, AWAITING_TTK_INGREDIENTS, AWAITING_COST_NAMES, AWAITING_COST_PRICES, AWAITING_ZAKUP_INPUT, AWAITING_SET_PRICE_BULK, AWAITING_DISHES_QTY = range(
    7)

# Заголовки
HEADERS = ["Ингредиент", "Вес (г)", "Цена за 1 кг", "Себестоимость"]
TTK_HEADERS = ["Блюдо", "Ингредиент", "Вес (г)"]

# =====================
# Утилиты Excel/матрицы
# =====================


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip()).lower()


def load_dishes_from_ttk(ttk_path: str = TTK_FILE,
                         sheet: str = TTK_SHEET_NAME) -> list[str]:
    """Читает список уникальных блюд из TTK.xlsx (лист TTK). Возвращает отсортированный список."""
    try:
        wb = load_workbook(ttk_path)
    except FileNotFoundError:
        return []
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    dishes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            name = str(row[0]).strip()
            if name:
                dishes.append(name)
    return sorted(set(dishes), key=str.lower)


def ensure_matrix_file(path: str):
    """Создать файл себестоимости с заголовками, если отсутствует. Вернуть (ws, wb)."""
    p = Path(path)
    if not p.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        ws.append(HEADERS)
        wb.save(p)
    wb = load_workbook(p)
    ws = wb.active
    # Проверим заголовки
    first_row = [c.value for c in ws[1]]
    if not first_row or first_row[:len(HEADERS)] != HEADERS:
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADERS)
        wb.save(p)
    return ws, wb


def get_or_create_ttk_ws(path: str = TTK_FILE,
                         sheet_name: str = TTK_SHEET_NAME):
    """Открыть TTK.xlsx и вернуть лист с точным именем sheet_name. Создать, если нет. Заголовки гарантированы."""
    p = Path(path)
    if p.exists():
        wb = load_workbook(p)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # удалить дефолтный Sheet при создании нового файла
        if len(wb.sheetnames) == 1 and wb.sheetnames[0].lower().startswith(
                "sheet"):
            del wb[wb.sheetnames[0]]
        ws = wb.create_sheet(title=sheet_name)

    # Заголовки
    if ws.max_row == 1 and ws.max_column == 1 and (ws["A1"].value is None):
        ws.append(TTK_HEADERS)
        wb.save(path)
    return ws, wb


def upsert_row(ws, name: str, weight_g: float, price_per_kg: float):
    """Добавить/обновить строку в матрице EXCEL_FILE."""
    target = normalize_name(name)
    found_row = None
    for row in ws.iter_rows(min_row=2):
        cell_name = row[0].value
        if cell_name and normalize_name(str(cell_name)) == target:
            found_row = row[0].row
            break
    if found_row:
        ws.cell(found_row, 2).value = weight_g
        ws.cell(found_row, 3).value = price_per_kg
    else:
        ws.append([name.strip(), weight_g, price_per_kg, ""])


def sort_sheet_by_name(ws):
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)]
    rows.sort(key=lambda r: normalize_name(str(r[0]) if r and r[0] else ""))
    ws.delete_rows(2, ws.max_row)
    for r in rows:
        ws.append(r)
    return ws


async def compute_purchase_for_dishes(
        dishes_portions: list[tuple[str, int]]) -> tuple[str | None, str]:
    """
    Возвращает (error, message). Если error is not None — это текст ошибки.
    Иначе message — готовый текст закупки (как в /zakup).
    """
    try:
        # Загрузка ТТК
        wb_ttk = load_workbook(TTK_FILE)
        if TTK_SHEET_NAME not in wb_ttk.sheetnames:
            return "❌ В TTK.xlsx нет листа с блюдами.", ""
        ws_ttk = wb_ttk[TTK_SHEET_NAME]

        # Загрузка матрицы
        if not Path(EXCEL_FILE).exists():
            return "❌ Файл с себестоимостью не найден.", ""
        df = pd.read_excel(EXCEL_FILE)

        result: dict[str, dict] = {}
        missing_products: list[str] = []

        for dish_name, portions in dishes_portions:
            found = False
            for row in ws_ttk.iter_rows(min_row=2, values_only=True):
                if row[0] and dish_name == str(row[0]).strip().lower():
                    ingredient = str(row[1]).strip()
                    weight = float(row[2])
                    total_weight = weight * portions

                    match = df[
                        df["Ингредиент"].str.lower().str.strip().str.contains(
                            ingredient.lower().strip(), na=False)]

                    if not match.empty and not pd.isna(
                            match.iloc[0]["Цена за 1 кг"]):
                        price_per_kg = float(match.iloc[0]["Цена за 1 кг"])
                        if ingredient in result:
                            result[ingredient]["weight"] += total_weight
                        else:
                            result[ingredient] = {
                                "weight": total_weight,
                                "price_per_kg": price_per_kg
                            }
                    else:
                        missing_products.append(ingredient)
                    found = True

            if not found:
                return f"❌ Блюдо «{dish_name}» не найдено в ТТК.", ""

        if missing_products:
            return f"❗️ Не найдены в матрице:\n{', '.join(sorted(set(missing_products)))}", ""

        # Формируем сообщение
        lines = []
        total_value = 0.0

        for ingredient, data in result.items():
            weight_grams = round(data["weight"])
            price_per_kg = float(data["price_per_kg"])
            cost = weight_grams / 1000 * price_per_kg
            total_value += cost
            lines.append(f"{ingredient} {weight_grams} гр")

        lines.append(f"\n💰 Общая сумма закупки: {round(total_value)} руб")
        return None, "\n".join(lines)

    except Exception as e:
        return f"⚠️ Ошибка при расчёте: {e}", ""


# =====================
# Парсинг для /ttк
# =====================

SUBS = {
    "тайский чили": "Соус сладкий чили",
    "креветка": "Креветки тигровые",
    "растительное масло": "Масло растительное",
}


def parse_ttk_multiline(text: str):
    """Парсит многострочный ввод: одна строка = "ингредиент вес". Возвращает [(name, grams), ...]"""
    rows: list[tuple[str, float]] = []
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line:
            continue
        m = re.match(r"(.+?)\s+(-?\d+[.,]?\d*)\s*$", line)
        if not m:
            continue
        name, qty = m.group(1).strip(), m.group(2).replace(",", ".")
        try:
            grams = float(qty)
        except ValueError:
            continue
        # Нормализация названий
        low = name.lower()
        for k, v in SUBS.items():
            if k in low:
                name = v
                break
        name = re.sub(r"\s+", " ", name).strip()
        if name:
            # Первая буква заглавная, остальное как есть
            name = name[0].upper() + name[1:]
            rows.append((name, grams))
    return rows


# =====================
# Команды бота
# =====================


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        ["➕ Новая ТТК", "📦 Закуп"],
        ["💰 Себестоимость", "🧾 Таблица"],
        ["🗑 Удалить", "📂 Все ТТК"],
        ["📑 Массовая ТТК"],
        ["🧮 Массовое обновление цен"],
        ["🧹 Очистить таблицу"],
        ["📊 Контроль маржи", "🔍 Отладка маржи"],
        ["🍽 Список блюд"],
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text("👨‍🍳 Бот запущен. Выберите действие:",
                                    reply_markup=reply_markup)


# -------- /dishes --------


async def dishes_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        await update.message.reply_text("⏳ Загружаю список блюд…")
        dishes = load_dishes_from_ttk()
        if not dishes:
            await update.message.reply_text(
                "❌ В TTK.xlsx не найдено ни одного блюда (лист TTK).")
            return ConversationHandler.END
        context.user_data["dishes_list"] = dishes
        context.user_data["cart"] = []
        context.user_data["dishes_page"] = 0
        await send_dishes_page(update, context, page=0)
        return AWAITING_DISHES_QTY
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка в /dishes: {e}")
        return ConversationHandler.END


async def send_dishes_page(update_or_cb, context, page: int):
    dishes = context.user_data.get("dishes_list", [])
    if not dishes:
        dishes = load_dishes_from_ttk()
        context.user_data["dishes_list"] = dishes
        if not dishes:
            if isinstance(update_or_cb, Update) and update_or_cb.message:
                await update_or_cb.message.reply_text("❌ В TTK.xlsx нет блюд.")
            else:
                await update_or_cb.callback_query.edit_message_text(
                    "❌ В TTK.xlsx нет блюд.")
            return

    total_pages = max(1, (len(dishes) + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))

    start = page * PAGE_SIZE
    end = start + PAGE_SIZE
    page_items = dishes[start:end]

    # Сохраняем текущую страницу
    context.user_data["dishes_page"] = page

    # Корзина
    cart = context.user_data.get("cart", [])
    cart_text = []
    if cart:
        cart_text.append(f"✅ Выбрано: {len(cart)} блюд")
        for i, dish in enumerate(cart[:5]):
            cart_text.append(f"{i+1}. {dish}")
        if len(cart) > 5:
            cart_text.append("…и ещё")
    else:
        cart_text.append("🛒 Корзина пуста")

    # Кнопки блюд
    rows = []
    for i, name in enumerate(page_items):
        global_idx = start + i  # индекс блюда в общем списке
        button_text = f"✓ {name}" if name in cart else name
        rows.append([
            InlineKeyboardButton(button_text,
                                 callback_data=f"dishes:pick:{global_idx}")
        ])

    # Навигация
    nav = []
    if page > 0:
        nav.append(
            InlineKeyboardButton("⬅️ Назад",
                                 callback_data=f"dishes:page:{page-1}"))
    if end < len(dishes):
        nav.append(
            InlineKeyboardButton("Вперёд ➡️",
                                 callback_data=f"dishes:page:{page+1}"))
    if nav:
        rows.append(nav)

    # Кнопки действий
    action_buttons = [
        InlineKeyboardButton("✅ Перейти к количествам",
                             callback_data="dishes:proceed"),
        InlineKeyboardButton("🗑 Очистить выбор", callback_data="dishes:clear")
    ]
    rows.append(action_buttons)

    kb = InlineKeyboardMarkup(rows)
    text = f"🍽 Выберите блюдо (страница {page+1}/{total_pages}):\n\n" + "\n".join(
        cart_text)
    if isinstance(update_or_cb, Update) and update_or_cb.message:
        await update_or_cb.message.reply_text(text, reply_markup=kb)
    else:
        query = update_or_cb.callback_query
        await query.edit_message_text(text, reply_markup=kb)


async def dishes_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    # Фоллбек: перечитаем блюда, если потерялись после рестарта
    if "dishes_list" not in context.user_data or not context.user_data.get(
            "dishes_list"):
        context.user_data["dishes_list"] = load_dishes_from_ttk()
        if not context.user_data["dishes_list"]:
            await query.edit_message_text("❌ В TTK.xlsx нет блюд.")
            return

    try:
        parts = data.split(":", 2)
        if len(parts) == 2:
            _, kind = parts
            payload = None
        elif len(parts) == 3:
            _, kind, payload = parts
        else:
            raise ValueError("Неверное количество частей в callback_data")
    except ValueError:
        await query.edit_message_text("⚠️ Неверный формат данных кнопки.")
        return

    if kind == "page":
        page = int(payload)
        await send_dishes_page(update, context, page)
        return

    if kind == "pick":
        try:
            idx = int(payload)
        except ValueError:
            await query.edit_message_text("⚠️ Ошибка индекса блюда.")
            return
        dishes = context.user_data["dishes_list"]
        if idx < 0 or idx >= len(dishes):
            await query.edit_message_text(
                "⚠️ Блюдо не найдено. Откройте список заново: /dishes")
            return
        dish = dishes[idx]

        cart = context.user_data.get("cart", [])
        if dish in cart:
            cart.remove(dish)
        else:
            cart.append(dish)
        context.user_data["cart"] = cart

        # Обновим страницу
        page = idx // PAGE_SIZE
        await send_dishes_page(update, context, page)
        return

    if kind == "clear":
        context.user_data["cart"] = []
        current_page = context.user_data.get("dishes_page", 0)
        await send_dishes_page(update, context, page=current_page)
        return

    if kind == "proceed":
        cart = context.user_data.get("cart", [])
        if not cart:
            await query.edit_message_text(
                "❌ Выберите хотя бы одно блюдо для продолжения.")
            return AWAITING_DISHES_QTY

        # Переходим к вводу количеств
        msg = "📝 Введите количество порций для каждого блюда (только числа в порядке списка):\n\n"
        for i, dish in enumerate(cart):
            msg += f"{i+1}. {dish}\n"
        msg += "\n✍️ Формат: числа через пробел, запятую или перенос строки. Например:\n"
        msg += "10 8 6\nили\n10,8,6\nили\n10\n8\n6"
        await query.edit_message_text(msg)
        return AWAITING_DISHES_QTY


async def handle_dishes_qty_input(update: Update,
                                  context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    cart = context.user_data.get("cart", [])
    # Разрешаем только числа и разделители; любые другие символы — мягкий репромпт
    text_only_digits = re.fullmatch(r"\s*\d+(?:[,\s]+\d+)*\s*", text)
    if not text_only_digits:
        await update.message.reply_text("Введите только количества цифрами (пример: 2 1 3).")
        return AWAITING_DISHES_QTY

    if not cart:
        await update.message.reply_text(
            "❌ Ошибка: корзина пуста. Начните заново: /dishes")
        return ConversationHandler.END

    # Разделяем строку на токены (числа) с учётом разных разделителей
    tokens = re.split(r'[,\s]+', text)
    numbers = []

    for token in tokens:
        if not token:
            continue
        try:
            num = int(token)
            if num <= 0:
                raise ValueError()
            numbers.append(num)
        except ValueError:
            await update.message.reply_text(
                f"❌ '{token}' не является положительным целым числом. Попробуйте снова."
            )
            return AWAITING_DISHES_QTY

    if len(numbers) != len(cart):
        msg = f"❌ Ожидалось {len(cart)} чисел для {len(cart)} блюд:\n"
        for i, dish in enumerate(cart):
            msg += f"{i+1}. {dish}\n"
        msg += f"А получено {len(numbers)}. Отправьте заново."
        await update.message.reply_text(msg)
        return AWAITING_DISHES_QTY

    # Формируем список пар (блюдо, количество)
    dishes_portions = [(dish.lower(), qty) for dish, qty in zip(cart, numbers)]

    # Рассчитываем закупку
    error, message = await compute_purchase_for_dishes(dishes_portions)
    if error:
        await update.message.reply_text(error)
        return AWAITING_DISHES_QTY

    await update.message.reply_text(message)
    return ConversationHandler.END


async def handle_text_buttons(update: Update,
                              context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    if text == "➕ Новая ТТК":
        return await update.message.reply_text("Введите команду /ttk")
    elif text == "📦 Закуп":
        return await update.message.reply_text("Введите команду /zakup")
    elif text == "💰 Себестоимость":
        return await update.message.reply_text("Введите команду /cost")
    elif text == "🧾 Таблица":
        return await send_table(update, context)
    elif text == "🗑 Удалить":
        await update.message.reply_text(
            "Введите команду в формате /delete <название позиции>")
    elif text == "📂 Все ТТК":
        return await send_ttk_file(update, context)
    elif text == "📑 Массовая ТТК":
        await update.message.reply_text("Введите команду /ttk")
    elif text == "🧮 Массовое обновление цен":
        return await start_set_price_bulk(update, context)
    elif text == "🧹 Очистить таблицу":
        return await clear_table(update, context)
    elif text == "📊 Контроль маржи":
        await update.message.reply_text("Введите команду /margin_check")
    elif text == "🔍 Отладка маржи":
        await update.message.reply_text("Введите команду /margin_debug")
    elif text == "🍽 Список блюд":
        # НЕ запускаем функцию напрямую, чтобы не обходить ConversationHandler
        await update.message.reply_text(
            "Открываю список блюд… Нажмите /dishes, если не появился.")
        return
    else:
        return await handle_text(update, context)


# ---------- /ttk ----------
async def start_ttk(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Введите название блюда:")
    return AWAITING_TTK_NAME


async def handle_ttk_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["ttk_dish_name"] = (update.message.text or "").strip()
    await update.message.reply_text(
        "Теперь введите состав блюда построчно (ингредиент вес):\n\nПример:\nКреветки тигровые 50\nКабачок 40\nСоус сладкий чили 20",
        reply_markup=ReplyKeyboardRemove(),
    )
    return AWAITING_TTK_INGREDIENTS


async def handle_ttk_ingredients(update: Update,
                                 context: ContextTypes.DEFAULT_TYPE):
    dish_name = context.user_data.get("ttk_dish_name", "Без названия")
    text = (update.message.text or "").strip()
    rows = parse_ttk_multiline(text)

    if not rows:
        await update.message.reply_text(
            "Не распознал ни одной позиции. Попробуйте ещё раз тем же форматом."
        )
        return AWAITING_TTK_INGREDIENTS

    try:
        ws, wb = get_or_create_ttk_ws(TTK_FILE, TTK_SHEET_NAME)
        for name, grams in rows:
            ws.append([dish_name, name, grams])
        wb.save(TTK_FILE)
    except Exception as e:
        await update.message.reply_text(f"⚠️ Ошибка при записи в TTK.xlsx: {e}"
                                        )
        return ConversationHandler.END

    await update.message.reply_text(
        f"✅ ТТК для блюда '{dish_name}' сохранён. Позиции: {len(rows)}.")
    return ConversationHandler.END


# ---------- /ttk_file ----------
async def send_ttk_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    p = Path(TTK_FILE)
    if not p.exists():
        await update.message.reply_text("❌ Файл TTK.xlsx не найден.")
        return
    try:
        wb = load_workbook(p)

        # если нет листа TTK — пытаемся использовать/переименовать Sheet1
        if TTK_SHEET_NAME not in wb.sheetnames:
            # 1) если есть единственный Sheet1 — переименуем и добавим заголовки при необходимости
            if "Sheet1" in wb.sheetnames and len(wb.sheetnames) == 1:
                ws = wb["Sheet1"]
                ws.title = TTK_SHEET_NAME
                if ws.max_row == 1 and ws.max_column == 1 and (ws["A1"].value
                                                               is None):
                    ws.append(TTK_HEADERS)
                wb.save(p)
            else:
                # 2) иначе просто создадим пустой TTK с заголовками
                ws = wb.create_sheet(title=TTK_SHEET_NAME)
                ws.append(TTK_HEADERS)
                wb.save(p)

        ws = wb[TTK_SHEET_NAME]
        dish_count = len(
            {row[0].value
             for row in ws.iter_rows(min_row=2) if row[0].value})

        with open(p, "rb") as f:
            await update.message.reply_document(f, filename="TTK.xlsx")
        await update.message.reply_text(
            f"✅ Файл TTK.xlsx отправлен.\n📝 Уникальных блюд: {dish_count}.")
    except Exception as e:
        await update.message.reply_text(f"⚠️ Ошибка обработки файла: {e}")


# ---------- /zakup ----------
async def start_zakup(update: Update,
                      context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "Введите название блюда и количество порций (каждая строка — отдельное блюдо):\n\n"
        "Пример:\nСтейк мачете 12\nСалат с хрустящими баклажанами 22",
        reply_markup=ReplyKeyboardRemove(),
    )
    return AWAITING_ZAKUP_INPUT


async def handle_zakup_input(update: Update,
                             context: ContextTypes.DEFAULT_TYPE):
    try:
        lines = (update.message.text or "").strip().split("\n")
        dishes: list[tuple[str, int]] = []
        missing_products: list[str] = []

        for line in lines:
            parts = line.rsplit(" ", 1)
            if len(parts) != 2:
                continue
            dish_name = parts[0].strip().lower()
            try:
                portions = int(parts[1])
            except ValueError:
                continue
            dishes.append((dish_name, portions))

        if not dishes:
            await update.message.reply_text("❌ Ошибка при обработке строки.")
            return ConversationHandler.END

        # Загрузка ТТК
        wb_ttk = load_workbook(TTK_FILE)
        if TTK_SHEET_NAME not in wb_ttk.sheetnames:
            await update.message.reply_text(
                f"❌ В TTK.xlsx нет листа '{TTK_SHEET_NAME}'.")
            return ConversationHandler.END
        ws_ttk = wb_ttk[TTK_SHEET_NAME]

        # Загрузка матрицы
        df = pd.read_excel(EXCEL_FILE)

        result: dict[str, dict] = {}

        for dish_name, portions in dishes:
            found = False
            for row in ws_ttk.iter_rows(min_row=2, values_only=True):
                if row[0] and dish_name == str(row[0]).strip().lower():
                    ingredient = str(row[1]).strip()
                    weight = float(row[2])
                    total_weight = weight * portions

                    match = df[
                        df["Ингредиент"].str.lower().str.strip().str.contains(
                            ingredient.lower().strip(), na=False)]

                    if not match.empty and not pd.isna(
                            match.iloc[0]["Цена за 1 кг"]):
                        price_per_kg = float(match.iloc[0]["Цена за 1 кг"])
                        if ingredient in result:
                            result[ingredient]["weight"] += total_weight
                        else:
                            result[ingredient] = {
                                "weight": total_weight,
                                "price_per_kg": price_per_kg
                            }
                    else:
                        missing_products.append(ingredient)
                    found = True

            if not found:
                await update.message.reply_text(
                    f"❌ Блюдо «{dish_name}» не найдено в ТТК.")
                return ConversationHandler.END

        if missing_products:
            await update.message.reply_text(
                f"❗️ Не найдены в матрице:\n{', '.join(sorted(set(missing_products)))}"
            )
            return ConversationHandler.END

        # Формируем сообщение
        lines = []
        total_value = 0.0

        for ingredient, data in result.items():
            weight_grams = round(data["weight"])
            price_per_kg = float(data["price_per_kg"])
            cost = weight_grams / 1000 * price_per_kg
            total_value += cost
            lines.append(f"{ingredient} {weight_grams} гр")

        lines.append(f"\n💰 Общая сумма закупки: {round(total_value)} руб")
        await update.message.reply_text("\n".join(lines))

    except Exception as e:
        await update.message.reply_text(f"⚠️ Ошибка: {e}")

    return ConversationHandler.END


# ---------- /cost ----------
async def start_cost(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Введите название блюда (или до 3 через Enter):",
        reply_markup=ReplyKeyboardRemove(),
    )
    return AWAITING_COST_NAMES


async def handle_cost_names(update: Update,
                            context: ContextTypes.DEFAULT_TYPE):
    dish_names = [
        name.strip().lower()
        for name in re.split(r"[\r\n]+", (update.message.text or "").strip())
        if name.strip()
    ]

    if not (1 <= len(dish_names) <= 3):
        await update.message.reply_text("⚠️ Введите от 1 до 3 названий блюд.")
        return AWAITING_COST_NAMES

    if not Path(EXCEL_FILE).exists() or not Path(TTK_FILE).exists():
        await update.message.reply_text(
            "❌ Не найден файл себестоимости или ТТК.")
        return ConversationHandler.END

    # Матрица цен (активный лист)
    wb_price = load_workbook(EXCEL_FILE)
    ws_price = wb_price.active

    headers = [cell.value for cell in ws_price[1]]
    name_col = headers.index("Ингредиент") if "Ингредиент" in headers else 0
    price_col = headers.index(
        "Цена за 1 кг") if "Цена за 1 кг" in headers else 2

    prices = {
        str(row[name_col].value).strip().lower(): row[price_col].value
        for row in ws_price.iter_rows(min_row=2)
        if row[name_col].value and row[price_col].value
    }

    # ТТК строго из листа TTK
    wb_ttk = load_workbook(TTK_FILE)
    if TTK_SHEET_NAME not in wb_ttk.sheetnames:
        await update.message.reply_text(
            f"❌ В TTK.xlsx нет листа '{TTK_SHEET_NAME}'.")
        return ConversationHandler.END
    ws_ttk = wb_ttk[TTK_SHEET_NAME]

    results = []
    lines = ["📊 Расчёт себестоимости:\n"]
    missing_global = set()

    for dish in dish_names:
        name = dish.strip().lower()
        rows = [
            r for r in ws_ttk.iter_rows(min_row=2)
            if str(r[0].value).strip().lower() == name
        ]

        if not rows:
            lines.append(f"❌ Блюдо «{dish}» не найдено в ТТК.")
            continue

        total = 0
        missing = []
        lines.append(f"🍽 *{dish.title()}*")

        for row in rows:
            ingr = str(row[1].value).strip().lower()
            weight = float(row[2].value)

            matched_price = None
            for k in prices:
                if ingr in k or k in ingr:
                    matched_price = prices[k]
                    break

            if matched_price is None:
                lines.append(f"— {ingr.title()} — ❌ нет цены")
                missing.append(ingr)
                missing_global.add(ingr)
                continue

            cost = round(weight * float(matched_price) / 1000)
            total += cost
            lines.append(
                f"— {ingr.title()}: {weight} г × {round(float(matched_price))}₽ = {cost}₽"
            )

        if missing:
            lines.append("⛔ Расчёт остановлен: не найдены все цены.\n")
            continue

        lines.append(f"💰 *Сумма:* {total}₽")
        results.append({"name": dish, "cost": total})

    if missing_global:
        lines.append("\n❗️Добавьте в матрицу цены для:")
        for ingr in sorted(missing_global):
            lines.append(f"— {ingr.title()}")
        await update.message.reply_text("\n".join(lines))
        return ConversationHandler.END

    context.user_data["cost_results"] = results
    await update.message.reply_text(
        "\n".join(lines) +
        "\n\n✍️ Введите цену продажи каждого блюда (с новой строки):")
    return AWAITING_COST_PRICES


async def handle_cost_price(update: Update,
                            context: ContextTypes.DEFAULT_TYPE):
    prices_text = (update.message.text or "").strip().split("\n")
    results = context.user_data.get("cost_results", [])

    if len(prices_text) != len(results):
        await update.message.reply_text(
            "⚠️ Количество цен не совпадает с количеством блюд.")
        return AWAITING_COST_PRICES

    response = ["📈 *Сводка:*"]
    total_cost = 0
    total_foodcost = 0

    for i, price_str in enumerate(prices_text):
        try:
            sale_price = float(price_str.strip().replace(",", "."))
        except ValueError:
            await update.message.reply_text(
                "⚠️ Введите корректные числовые значения.")
            return AWAITING_COST_PRICES

        dish = results[i]["name"]
        cost = results[i]["cost"]
        foodcost = round(cost / sale_price * 100, 1) if sale_price else 0

        total_cost += cost
        total_foodcost += foodcost

        response.append(
            f"• {dish.title()} — себестоимость {cost}₽, продажа {sale_price}₽, foodcost {foodcost}%"
        )

    avg_foodcost = round(total_foodcost / len(results), 1) if results else 0

    response.append(f"\n💵 *Общая себестоимость:* {total_cost}₽")
    response.append(f"📊 *Средний foodcost:* {avg_foodcost}%")

    await update.message.reply_text("\n".join(response), parse_mode="Markdown")
    return ConversationHandler.END


# ---------- /delete ----------
async def delete_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = " ".join(context.args).strip().lower()
    deleted = False

    for file, col in [(EXCEL_FILE, 0), (TTK_FILE, 0)]:
        if not Path(file).exists():
            continue
        wb = load_workbook(file)
        # В TTK удаляем только на листе TTK
        if file == TTK_FILE:
            if TTK_SHEET_NAME not in wb.sheetnames:
                continue
            ws = wb[TTK_SHEET_NAME]
        else:
            ws = wb.active

        rows_to_delete = [
            row[0].row for row in ws.iter_rows(min_row=2)
            if str(row[col].value).strip().lower() == name
        ]
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)
            deleted = True
        wb.save(file)

    if deleted:
        await update.message.reply_text(f"✅ '{name}' удалён.")
    else:
        await update.message.reply_text(f"❌ '{name}' не найден.")


# ---------- /table ----------
async def send_table(update: Update, context: ContextTypes.DEFAULT_TYPE):
    p = Path(EXCEL_FILE)
    if not p.exists():
        await update.message.reply_text("❌ Файл не найден.")
        return
    with open(p, "rb") as f:
        await update.message.reply_document(f, filename="Себестоимость.xlsx")


# ---------- /set_price_bulk ----------


def clean_price_bulk_input(text: str) -> list[str]:
    text = text.replace("\xa0", " ")
    lines = text.split("\n")
    cleaned: list[str] = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        line = re.sub(r"[^\w\s.,-]", "", line)
        line = line.replace(",", ".")
        line = re.sub(r"\s+", " ", line)
        cleaned.append(line)
    return cleaned


async def start_set_price_bulk(update: Update,
                               context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "Введите список продуктов, цен и веса (каждый с новой строки):\n\n"
        "Пример:\nБрокколи 4000 250\nКунжут 1000 320\nСоус песто 500 280",
        reply_markup=ReplyKeyboardRemove(),
    )
    return AWAITING_SET_PRICE_BULK


async def handle_set_price_bulk(update: Update,
                                context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    cleaned_lines = clean_price_bulk_input(text)
    ws, wb = ensure_matrix_file(EXCEL_FILE)
    added = updated = 0
    bad: list[str] = []
    updated_items: list[tuple[str, float]] = []

    for line in cleaned_lines:
        parts = line.rsplit(" ", 2)
        if len(parts) != 3:
            bad.append(line)
            continue
        name, w_str, s_str = parts
        w_num = re.sub(r"\s", "", w_str)
        s_num = re.sub(r"\s", "", s_str)
        try:
            weight_g = float(w_num)
            amount = float(s_num)
            if weight_g <= 0 or amount <= 0:
                raise ValueError()
            price_per_kg = amount / weight_g * 1000.0
        except Exception:
            bad.append(line)
            continue
        before = ws.max_row
        upsert_row(ws, name, weight_g, round(price_per_kg, 2))
        after = ws.max_row
        if after > before:
            added += 1
        else:
            updated += 1
        updated_items.append((name, round(price_per_kg, 2)))

    sort_sheet_by_name(ws)
    wb.save(EXCEL_FILE)

    msg = [f"✅ Готово. Добавлено: {added}, обновлено: {updated}."]
    if updated_items:
        msg.append("\nОбновлённые позиции:")
        for name, price_per_kg in updated_items[:30]:
            msg.append(f"— {name}: {price_per_kg} ₽/кг")
    if bad:
        msg.append("\n⚠️ Не распарсились строки:")
        msg.extend([f"— {b}" for b in bad[:20]])
        if len(bad) > 20:
            msg.append(f"…и ещё {len(bad) - 20}")

    await update.message.reply_text("\n".join(msg))
    return ConversationHandler.END


# ---------- /clear_table ----------
async def clear_table(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ws, wb = ensure_matrix_file(EXCEL_FILE)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    wb.save(EXCEL_FILE)
    await update.message.reply_text(
        "🧹 Таблица очищена. Оставлены только заголовки.")


# ---------- Обработка одиночной строки ввода (не список) ----------
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text or ""

    # Если это список — подскажем использовать пакетную загрузку
    if "\n" in raw.strip():
        await update.message.reply_text(
            "Похоже, вы отправили список из нескольких строк. Для пакетной загрузки используйте команду /set_price_bulk."
        )
        return

    text = raw.strip().lower()
    text = (text.replace("грамм", "").replace("грамма",
                                              "").replace("гр", "").replace(
                                                  "руб", "").replace("₽", ""))
    text = re.sub(r"\s+", " ", text).strip()

    m = re.match(r"^(.+?)\s+([\d.,]+)\s+([\d.,]+)$", text)
    if not m:
        await update.message.reply_text(
            "Не удалось распознать строку. Формат: «Название 1000 250».\nДля списка позиций используйте /set_price_bulk."
        )
        return

    name, w_str, p_str = m.groups()
    try:
        weight = float(w_str.replace(",", "."))
        amount = float(p_str.replace(",", "."))
        if weight <= 0 or amount <= 0:
            raise ValueError()
        price_per_kg = round(amount / weight * 1000.0, 2)
    except Exception:
        await update.message.reply_text(
            "Не удалось прочитать вес/цену. Пример: «Брокколи 1000 250»")
        return

    try:
        ws, wb = ensure_matrix_file(EXCEL_FILE)
        upsert_row(ws, name, weight, price_per_kg)
        sort_sheet_by_name(ws)
        wb.save(EXCEL_FILE)
        await update.message.reply_text(
            f"✅ Записано: {name.strip()} — {int(weight)} г, {amount:g} ₽, {price_per_kg} ₽/кг"
        )
    except Exception:
        await update.message.reply_text("❌ Ошибка при обработке строки.")


# =====================
# Запуск
# =====================


def main():
    from dotenv import load_dotenv
    load_dotenv()

    application = ApplicationBuilder().token(os.getenv("BOT_TOKEN")).post_init(set_bot_commands).build()

    # Подключение модуля контроля маржи
    setup_margin_control(application)

    # Команды
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("ttk_file", send_ttk_file))
    application.add_handler(CommandHandler("table", send_table))
    application.add_handler(CommandHandler("delete", delete_entry))
    application.add_handler(CommandHandler("clear_table", clear_table))

    # /ttk
    conv_ttk = ConversationHandler(
        entry_points=[CommandHandler("ttk", start_ttk),
        MessageHandler(filters.Regex(r"^➕ Новая ТТК$"), start_ttk)],
        states={
            AWAITING_TTK_NAME:
            [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_ttk_name)],
            AWAITING_TTK_INGREDIENTS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND,
                               handle_ttk_ingredients)
            ],
        },
        fallbacks=[],
    )
    application.add_handler(conv_ttk)

    # /cost
    conv_cost = ConversationHandler(
        entry_points=[CommandHandler("cost", start_cost),
        MessageHandler(filters.Regex(r"^💰 Себестоимость$"), start_cost)],
        states={
            AWAITING_COST_NAMES: [
                MessageHandler(filters.TEXT & ~filters.COMMAND,
                               handle_cost_names)
            ],
            AWAITING_COST_PRICES: [
                MessageHandler(filters.TEXT & ~filters.COMMAND,
                               handle_cost_price)
            ],
        },
        fallbacks=[],
    )
    application.add_handler(conv_cost)

    # /zakup
    conv_zakup = ConversationHandler(
        entry_points=[CommandHandler("zakup", start_zakup),
        MessageHandler(filters.Regex(r"^📦 Закуп$"), start_zakup)],
        states={
            AWAITING_ZAKUP_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND,
                               handle_zakup_input)
            ],
        },
        fallbacks=[],
    )
    application.add_handler(conv_zakup)

    # /set_price_bulk
    conv_set_price_bulk = ConversationHandler(
        entry_points=[CommandHandler("set_price_bulk", start_set_price_bulk)],
        states={
            AWAITING_SET_PRICE_BULK: [
                MessageHandler(filters.TEXT & ~filters.COMMAND,
                               handle_set_price_bulk)
            ],
        },
        fallbacks=[],
    )
    application.add_handler(conv_set_price_bulk)

    # /dishes (разговор и inline-кнопки внутри состояния)
    conv_dishes = ConversationHandler(
        entry_points=[
            CommandHandler("dishes", dishes_command),
            MessageHandler(filters.Regex(r"^🍽 Список блюд$"), dishes_command),
        ],
        states={
            AWAITING_DISHES_QTY: [
                CallbackQueryHandler(dishes_callback, pattern=r"^dishes:"),
                MessageHandler(filters.TEXT & ~filters.COMMAND,
                               handle_dishes_qty_input),
            ],
        },
        fallbacks=[],
    )
    application.add_handler(conv_dishes)

    # Текстовые кнопки и прочий текст (последним)
    application.add_handler(
        MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_buttons))

    print("✅ Бот запущен...")
    application.run_polling()


if __name__ == "__main__":
    main()
